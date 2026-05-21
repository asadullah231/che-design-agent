"""
Export simulation results to Excel and PDF
"""

import os
from datetime import datetime


EXPORTS_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "exports")


class ResultExporter:

    def __init__(self):
        os.makedirs(EXPORTS_DIR, exist_ok=True)

    def to_excel(self, results: dict, filename: str) -> str:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.chart import BarChart, Reference

            if not filename.endswith(".xlsx"):
                filename += ".xlsx"
            filepath = os.path.join(EXPORTS_DIR, filename)

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Design Results"

            # Header styling
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_fill = PatternFill("solid", fgColor="1F4E79")
            center = Alignment(horizontal="center")
            thin = Side(style="thin", color="000000")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            # Title
            ws.merge_cells("A1:C1")
            ws["A1"] = "Chemical Engineering Design Report"
            ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
            ws["A1"].alignment = center

            ws.merge_cells("A2:C2")
            ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
            ws["A2"].alignment = center

            # Column headers
            ws["A4"] = "Parameter"
            ws["B4"] = "Value"
            ws["C4"] = "Unit"
            for cell in [ws["A4"], ws["B4"], ws["C4"]]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center
                cell.border = border

            # Data rows
            unit_map = {
                "feed_flowrate_kghr": "kg/hr",
                "distillate_flowrate_kghr": "kg/hr",
                "bottoms_flowrate_kghr": "kg/hr",
                "column_diameter_m": "m",
                "column_height_m": "m",
                "condenser_duty_kW": "kW",
                "reboiler_duty_kW": "kW",
                "operating_pressure_kPa": "kPa",
                "heat_duty_kW": "kW",
                "heat_transfer_area_m2": "m²",
                "LMTD_C": "°C",
                "U_overall_W_m2K": "W/m².K",
            }

            row = 5
            alt_fill = PatternFill("solid", fgColor="D6E4F0")
            for key, value in results.items():
                if isinstance(value, (dict, list)):
                    continue
                ws.cell(row=row, column=1, value=self._format_key(key)).border = border
                ws.cell(row=row, column=2, value=value).border = border
                ws.cell(row=row, column=3, value=unit_map.get(key, "-")).border = border
                if row % 2 == 0:
                    for col in range(1, 4):
                        ws.cell(row=row, column=col).fill = alt_fill
                row += 1

            ws.column_dimensions["A"].width = 35
            ws.column_dimensions["B"].width = 20
            ws.column_dimensions["C"].width = 15

            wb.save(filepath)
            return filepath

        except ImportError:
            return self._to_excel_csv_fallback(results, filename)

    def _to_excel_csv_fallback(self, results: dict, filename: str) -> str:
        csv_name = filename.replace(".xlsx", ".csv")
        filepath = os.path.join(EXPORTS_DIR, csv_name)
        with open(filepath, "w") as f:
            f.write("Parameter,Value\n")
            for k, v in results.items():
                if not isinstance(v, (dict, list)):
                    f.write(f"{k},{v}\n")
        return filepath

    def to_pdf(self, results: dict, filename: str, engineer: str = "Engineer", project: str = "Project") -> str:
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import cm
            from reportlab.lib import colors
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer
            from reportlab.lib.enums import TA_CENTER

            if not filename.endswith(".pdf"):
                filename += ".pdf"
            filepath = os.path.join(EXPORTS_DIR, filename)

            doc = SimpleDocTemplate(filepath, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm)
            styles = getSampleStyleSheet()
            story = []

            # Title
            title_style = ParagraphStyle("title", parent=styles["Title"],
                                         textColor=colors.HexColor("#1F4E79"), fontSize=18)
            story.append(Paragraph("Chemical Engineering Design Report", title_style))
            story.append(Spacer(1, 0.5*cm))
            story.append(Paragraph(f"Engineer: {engineer} | Project: {project}", styles["Normal"]))
            story.append(Paragraph(f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles["Normal"]))
            story.append(Spacer(1, 1*cm))

            # Results table
            table_data = [["Parameter", "Value"]]
            for key, value in results.items():
                if not isinstance(value, (dict, list)):
                    table_data.append([self._format_key(key), str(value)])

            table = Table(table_data, colWidths=[10*cm, 7*cm])
            table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 12),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#D6E4F0")]),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 1), (-1, -1), 10),
            ]))
            story.append(table)

            doc.build(story)
            return filepath

        except ImportError:
            return self._to_pdf_text_fallback(results, filename, engineer, project)

    def _to_pdf_text_fallback(self, results, filename, engineer, project) -> str:
        txt_name = filename.replace(".pdf", "_report.txt")
        filepath = os.path.join(EXPORTS_DIR, txt_name)
        with open(filepath, "w") as f:
            f.write(f"CHEMICAL ENGINEERING DESIGN REPORT\n")
            f.write(f"Engineer: {engineer} | Project: {project}\n")
            f.write(f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}\n")
            f.write("=" * 50 + "\n\n")
            for k, v in results.items():
                if not isinstance(v, (dict, list)):
                    f.write(f"{self._format_key(k)}: {v}\n")
        return filepath

    def _format_key(self, key: str) -> str:
        return key.replace("_", " ").title()
